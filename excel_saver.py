import os
import logging
from datetime import datetime
import openpyxl
from threading import Lock

# Logger
logger = logging.getLogger("excel_saver")
logger.setLevel(logging.INFO)

excel_lock = Lock()
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, 'src', 'LOS', 'Feedback-training.xlsx')

def append_feedback_to_excel(data: dict):
    logger.info("append_feedback_to_excel: keys=%s", list(data.keys()))
    with excel_lock:
        try:
            if not os.path.exists(EXCEL_PATH):
                os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Свод"
                ws.append([
                    "Партнер", "Дата", "Фамилия и имя спикера/ои",
                    "Положительные качества", "Отрицательные",
                    "Полезность информации", "Аргументация",
                    "Самые яркие мысли с обучения", "Что добавить в тренинг",
                    "Если выбрана статистика", "Общее впечатление",
                    "Настроение", "NPS"
                ])
                wb.save(EXCEL_PATH)
                wb.close()
                logger.info("Created new Excel file %s", EXCEL_PATH)

            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb["Свод"] if "Свод" in wb.sheetnames else wb.create_sheet("Свод")

            partner = data.get("partner", "")
            dt = datetime.fromisoformat(data.get("dateTime"))
            date_str = dt.strftime("%d.%m.%Y %H:%M")

            speakers = data.get("speakersFeedback", [])
            names = [sp.get("fullName","") for sp in speakers if sp.get("fullName")]
            col_c = ", ".join(names) if names else "-"

            qualities = []
            for sp in speakers:
                qualities += [str(q) for q in sp.get("qualities", []) if q is not None]
            col_d = ", ".join(qualities) if qualities else "-"

            ca = data.get("commonAnswers", {})
            col_f = ca.get("usefulness","")
            col_g = ca.get("uselessArgument","")
            col_h = ca.get("brightThoughts","")
            col_i = ca.get("additionalSuggestions","")
            col_j = ca.get("statsDetails","")
            col_k = ca.get("impression","")
            col_l = ca.get("mood","")
            col_m = ca.get("recommendation","")

            ws.append([
                partner, date_str, col_c, col_d, "-",
                col_f, col_g, col_h, col_i, col_j,
                col_k, col_l, col_m
            ])

            wb.save(EXCEL_PATH)
            wb.close()
            logger.info("Appended new row to %s", EXCEL_PATH)

        except Exception:
            logger.exception("Error in append_feedback_to_excel")



