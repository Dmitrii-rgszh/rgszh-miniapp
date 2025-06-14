import os
import json
import openpyxl
import requests
import smtplib
import pythoncom
import threading
from threading import Lock
from datetime import datetime

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from win32com.client import DispatchEx, constants

XLUP = -4162

# создаём приложение
app = Flask(__name__, static_folder='build', static_url_path='')

# список доменов, с которых разрешаем запросы
app = Flask(__name__, static_folder='build', static_url_path='')
CORS(app, resources={r"/api/*": {"origins": "*"}})
excel_lock = Lock()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, 'src', 'LOS', 'Feedback-training.xlsx')

# Утилита для CORS preflight
def cors_preflight_response():
    origin = request.headers.get('Origin', allowed_origins[-1])
    resp = jsonify({'message': 'CORS preflight passed'})
    if origin in allowed_origins:
        resp.headers.update({
            'Access-Control-Allow-Origin': origin,
            'Access-Control-Allow-Headers': 'Content-Type, Authorization',
            'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
            'Access-Control-Allow-Credentials': 'true'
        })
    return resp, 204


@app.after_request
def after_request(response):
    origin = request.headers.get('Origin', allowed_origins[-1])
    if origin in allowed_origins:
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    return response


# ===== Эндпоинты POLL (без изменений) =====

poll_lock = Lock()
poll_results = [
    {"text": "Чтобы отстал руководитель", "votes": 0},
    {"text": "Чтобы перевыполнить план и заработать много деняк", "votes": 0},
    {"text": "Чтобы победить в конкурсе и полетать на самолёте", "votes": 0},
    {"text": "Чтобы клиент меня любил сильнее, чем маму", "votes": 0},
    {"text": "Какого ещё маржа? Я не в курсе", "votes": 0}
]

@app.route('/api/poll', methods=["GET"])
def get_poll():
    with poll_lock:
        return jsonify(poll_results)

@app.route('/api/poll/vote', methods=["POST"])
def vote_poll():
    data = request.get_json()
    idx = data.get('index')
    if idx is None or not isinstance(idx, int) or idx < 0 or idx >= len(poll_results):
        return jsonify({"error": "Invalid index"}), 400
    with poll_lock:
        poll_results[idx]["votes"] += 1
        socketio.emit('pollResults', poll_results)
    return jsonify({"success": True})

@app.route('/api/poll/reset', methods=["POST"])
def reset_poll():
    with poll_lock:
        for opt in poll_results:
            opt["votes"] = 0
        socketio.emit('pollResults', poll_results)
    return jsonify({"success": True})


# ===== Эндпоинты FEEDBACK =====
@app.route('/api/feedback/save', methods=["POST", "OPTIONS"])
def save_feedback():
    # для preflight-запроса
    if request.method == "OPTIONS":
        return '', 200

    payload = request.get_json(force=True)
    # Стартуем фоновой поток и тут же отдаем 200
    threading.Thread(
        target=_write_feedback_to_excel,
        args=(payload,),
        daemon=True
    ).start()

    return jsonify({"status": "queued"}), 200

@app.route('/api/proxy/feedback/save', methods=["POST", "OPTIONS"])
def proxy_feedback():
    if request.method == "OPTIONS":
        return cors_preflight_response()
    data = request.get_json(force=True)
    try:
        url = 'https://rgszh-miniapp.org/api/feedback/save'
        resp = requests.post(url, json=data, headers={'Content-Type': 'application/json'})
        return Response(
            resp.content,
            status=resp.status_code,
            content_type=resp.headers.get('Content-Type')
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def _write_feedback_to_excel(data: dict):
    with excel_lock:
        # Если файл не существует — создаём и пишем заголовки
        if not os.path.exists(EXCEL_PATH):
            os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Свод"
            headers = ["ДатаВремя", "Партнёр", "ОбщаяИнформация", "Спикеры", "NPS_рекомендация"]
            ws.append(headers)
            wb.save(EXCEL_PATH)

        # Загружаем книгу и лист "Свод"
        wb = openpyxl.load_workbook(EXCEL_PATH)
        # Если листа «Свод» нет — создаём
        if "Свод" in wb.sheetnames:
            ws = wb["Свод"]
        else:
            ws = wb.create_sheet("Свод")
            headers = ["ДатаВремя", "Партнёр", "ОбщаяИнформация", "Спикеры", "NPS_рекомендация"]
            ws.append(headers)

        # Формируем строку и дописываем
        row = [
            datetime.now().isoformat(),
            data.get("partner", ""),
            json.dumps(data.get("commonAnswers", {}), ensure_ascii=False),
            json.dumps(data.get("speakersFeedback", []), ensure_ascii=False),
            data.get("commonAnswers", {}).get("recommendation", "")
        ]
        ws.append(row)

        # Сохраняем изменения
        wb.save(EXCEL_PATH)
        wb.close()

        print(f"[Excel] Row appended to {EXCEL_PATH} at {row[0]}")


# ===== Эндпоинты SNP (расчёт через Excel) =====

@app.route('/api/snp/save', methods=["POST", "OPTIONS"])
def snp_save():
    if request.method == "OPTIONS":
        return cors_preflight_response()
    data = request.get_json()
    print("[Server] SNP расчёт получен:", data)
    try:
        pythoncom.CoInitialize()
        excel = DispatchEx("Excel.Application")
        try:
            excel.Visible = False
        except:
            pass

        excel_path = r"C:\Users\shapeless\Desktop\miniapp\src\calculators\СНП.xlsb"
        wb = excel.Workbooks.Open(excel_path, ReadOnly=0)
        ws = wb.Sheets("Информация о клиенте")

        ws.Range("I16").Value = data.get("payment", "")
        ws.Range("I17").Value = data.get("sum", 0)
        ws.Range("I23").Value = data.get("gender", "")
        ws.Range("I24").Value = data.get("birthdate", "")

        # Полный пересчёт книги
        excel.CalculateFull()
        print("[Server] SNP: выполнен excel.CalculateFull()")
        wb.Save()

        ws_res = wb.Sheets("Предложение для клиента")
        def rv(r):
            v = ws_res.Range(r).Value
            return v if (v is not None) else 0

        result = {
            "calculationDate":   datetime.now().strftime("%d.%m.%Y"),
            "gender":            data.get("gender",""),
            "age":               data.get("age",0),
            "payment":           data.get("payment","").lower(),
            "basePremium":       rv("F20"),
            "totalPremium":      rv("H41"),
            "termYears":         rv("K9"),
            "frequency":         rv("K10"),
            "contribution":      rv("C13"),
            "totalContribution": rv("C16"),
            "accrualPercent":    rv("H12"),
            "accrualAmount":     rv("H13"),
            "totalAccrual":      rv("H16"),
            "finalAnnual":       rv("M13"),
            "finalTotal":        rv("M16")
        }

        wb.Close(SaveChanges=True)
        excel.Quit()
        pythoncom.CoUninitialize()

        return jsonify(result), 200

    except Exception as e:
        print("[Server] Ошибка SNP:", e)
        try:
            excel.Quit()
            pythoncom.CoUninitialize()
        except:
            pass
        return jsonify({"error": str(e)}), 500

@app.route('/api/proxy/snp/save', methods=["POST", "OPTIONS"])
def proxy_snp_save():
    if request.method == "OPTIONS":
        return cors_preflight_response()
    data = request.get_json(force=True)
    try:
        target = url_for('snp_save', _external=True)
        resp = requests.post(target, json=data, headers={'Content-Type': 'application/json'})
        return Response(
            resp.content,
            status=resp.status_code,
            content_type=resp.headers.get('Content-Type')
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/assessment/candidate/save', methods=["POST", "OPTIONS"])
def save_assessment():
    """
    Сохраняет результаты опроса кандидата в файл assesment_candidates.xlsx.
    Ожидает JSON:
    {
      "surname": "...",
      "firstName": "...",
      "patronymic": "...",
      "dateTime": "...",
      "answers": ["...", "...", ...]
    }
    """
    if request.method == "OPTIONS":
        return cors_preflight_response()

    data = request.get_json(force=True)
    print("[Server] Получены данные оценки кандидата:", data)

    # Путь к вашему файлу Excel (лежит в корне проекта рядом с server.py)
    excel_path = os.path.join(app.root_path, 'assesment_candidates.xlsx')

    with excel_lock:
        # Инициализируем COM-библиотеки в этом потоке
        pythoncom.CoInitialize()

        # Если файл существует — открываем, иначе — создаём новый
        if os.path.exists(excel_path):
            excel = DispatchEx('Excel.Application')
            workbook = excel.Workbooks.Open(excel_path)
        else:
            # Создаём новый экземпляр Excel и новую книгу
            excel = DispatchEx('Excel.Application')
            workbook = excel.Workbooks.Add()
            worksheet = workbook.Worksheets(1)
            worksheet.Name = "Assessment"

            # Формируем заголовки: Фамилия, Имя, Отчество, ДатаВремя, затем Вопрос 1, Вопрос 2 и т.д.
            headers = ["Фамилия", "Имя", "Отчество", "ДатаВремя"] + [
                f"Вопрос {i+1}" for i in range(len(data.get('answers', [])))
            ]
            for col, header in enumerate(headers, start=1):
                worksheet.Cells(1, col).Value = header

            # Сохраняем новый файл под именем assesment_candidates.xlsx
            workbook.SaveAs(excel_path)
            workbook.Close(False)
            workbook = excel.Workbooks.Open(excel_path)

        worksheet = workbook.Worksheets(1)

        # Находим последнюю заполненную строку в столбце A (Фамилия)
        last_row = worksheet.Cells(worksheet.Rows.Count, 1).End(constants.xlUp).Row
        next_row = last_row + 1

        # Извлекаем данные из JSON
        surname = data.get('surname', '')
        first_name = data.get('firstName', '')
        patronymic = data.get('patronymic', '')
        date_time = data.get('dateTime', '')
        answers = data.get('answers', [])

        # Записываем ФИО и дату/время в столбцы 1–4
        worksheet.Cells(next_row, 1).Value = surname
        worksheet.Cells(next_row, 2).Value = first_name
        worksheet.Cells(next_row, 3).Value = patronymic
        worksheet.Cells(next_row, 4).Value = date_time

        # Записываем ответы, начиная с колонки 5
        for idx, answer in enumerate(answers):
            worksheet.Cells(next_row, 5 + idx).Value = answer

        # Сохраняем и закрываем книгу
        workbook.Save()
        workbook.Close(False)
        excel.Quit()
        pythoncom.CoUninitialize()

    return jsonify({"success": True}), 200


@app.route('/api/proxy/assessment/candidate/save', methods=["POST", "OPTIONS"])
def proxy_assessment_save():
    if request.method == "OPTIONS":
        return cors_preflight_response()

    data = request.get_json(force=True)
    try:
        target = url_for('save_assessment', _external=True)
        resp = requests.post(
            target,
            json=data,
            headers={'Content-Type': 'application/json'}
        )
        return Response(
            resp.content,
            status=resp.status_code,
            content_type=resp.headers.get('Content-Type')
        )
    except Exception as e:
        print("[Server] Ошибка proxy_assessment_save:", e)
        return jsonify({"error": str(e)}), 500

@app.route('/api/snp/send_manager', methods=["POST", "OPTIONS"])
def snp_send_manager():
    if request.method == "OPTIONS":
        return cors_preflight_response()
    data = request.get_json()
    try:
        subject = data.get("subject", "")
        body    = data.get("body", "")

        smtp_server = "smtp.yandex.ru"
        smtp_port   = 587
        smtp_user   = "rgszh-miniapp@yandex.ru"
        smtp_pass   = "ofuhaiuymneawcve"
        recipients  = ["zerotlt@mail.ru", "I.dav@mail.ru"]

        message = (
            f"From: {smtp_user}\r\n"
            f"To: {', '.join(recipients)}\r\n"
            f"Subject: {subject}\r\n"
            "Content-Type: text/plain; charset=utf-8\r\n\r\n"
            f"{body}"
        )

        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_user, recipients, message.encode('utf-8'))
        server.quit()

        return jsonify({"status": "ok"}), 200
    except Exception as e:
        print("[Server] Ошибка при отправке менеджеру SNP:", e)
        return jsonify({"error": str(e)}), 500

@app.route('/api/proxy/snp/send_manager', methods=["POST", "OPTIONS"])
def proxy_snp_send_manager():
    if request.method == "OPTIONS":
        return cors_preflight_response()
    data = request.get_json(force=True)
    try:
        target = url_for('snp_send_manager', _external=True)
        resp = requests.post(target, json=data, headers={'Content-Type': 'application/json'})
        return Response(
            resp.content,
            status=resp.status_code,
            content_type=resp.headers.get('Content-Type')
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ===== Эндпоинты CAREFUTURE (расчёт через файл «Забота о будущем.xlsb») =====

@app.route('/api/carefuture/save', methods=["POST", "OPTIONS"])
def carefuture_save():
    """
    Принимает JSON (POST):
    {
      "email": "...@vtb.ru",
      "birthdate": "DD.MM.YYYY",
      "age": <число>,
      "gender": "мужской" или "женский",
      "term": <число лет>,
      "payment": "Страхового взноса" или "Страховой суммы",
      "sum": <число: ежегодный взнос>
    }

    Возвращает JSON:
    {
      "contribution": <annual_sum>,
      "totalContribution": <annual_sum * term>,
      "accrualPercent": <ячейка H12>,
      "accrualAmount": <ячейка H13>,
      "totalAccrual": <ячейка H16>,
      "finalAnnual": <ячейка M13>,
      "finalTotal": <ячейка M16>,
      "accumulation": <ячейка G13>,
      "totalDeduction": <ячейка C22 (целое число)>,
      "delayedPayment": <ячейка G14>
    }
    """
    if request.method == "OPTIONS":
        return cors_preflight_response()

    data = request.get_json()
    term        = int(data.get("term", 0))
    payment     = data.get("payment", "")
    annual_sum  = int(data.get("sum", 0))  # ежегодный взнос

    try:
        pythoncom.CoInitialize()
        excel = DispatchEx("Excel.Application")
        try:
            excel.Visible = False
            excel.DisplayAlerts = False  # отключаем диалоги Excel
        except:
            pass

        excel_path = r"C:\Users\shapeless\Desktop\miniapp\src\calculators\Забота о будущем.xlsb"
        print(f"[Server] CAREFUTURE: Открываем: {excel_path}")
        wb = excel.Workbooks.Open(excel_path, ReadOnly=0)
        print("[Server] CAREFUTURE: Excel открыт")

        ws = wb.Sheets("Информация о клиенте")
        print("[Server] CAREFUTURE: Лист 'Информация о клиенте' найден")

        # 1) Вставляем срок программы (число лет) в I15
        ws.Range("I15").Value = term
        print(f"[Server] CAREFUTURE: I15 ← {term}")

        # 2) Вставляем сумму ежегодного взноса
        print(f"[Server] CAREFUTURE: payment='{payment}', sum={annual_sum}")
        if payment == "Страхового взноса":
            ws.Range("I17").Value = annual_sum
            ws.Range("I18").Value = 0
            print(f"[Server] CAREFUTURE: I17 ← {annual_sum}")
        else:
            ws.Range("I18").Value = annual_sum
            ws.Range("I17").Value = 0
            print(f"[Server] CAREFUTURE: I18 ← {annual_sum}")

        # 3) Вставляем дату рождения в I23
        birthdate_str = data.get("birthdate", "")
        ws.Range("I23").Value = birthdate_str
        print(f"[Server] CAREFUTURE: I23 ← {birthdate_str}")

        # Вместо полного пересчёта вызываем только локальный Calculate() на активном листе
        try:
            ws.Calculate()
            print("[Server] CAREFUTURE: Выполнен ws.Calculate()")
        except Exception as e_calc:
            print("[Server] CAREFUTURE: ОШИБКА при ws.Calculate():", e_calc)

        # Не сохраняем книгу (SaveChanges=False) — чтобы не блокировать COM-интерфейс
        # Сразу переходим к чтению результатов с листа "Предложение для клиента"
        try:
            ws_res = wb.Sheets("Предложение для клиента")
            print("[Server] CAREFUTURE: Лист 'Предложение для клиента' найден")
        except Exception as e_sheet:
            print("[Server] CAREFUTURE: НЕ НАЙДЕН лист 'Предложение для клиента':", e_sheet)
            wb.Close(SaveChanges=False)
            excel.Quit()
            pythoncom.CoUninitialize()
            return jsonify({"error": "Лист 'Предложение для клиента' не найден"}), 500

        def rv(addr):
            v = ws_res.Range(addr).Value
            return v if (v is not None) else 0

        accrual_percent = rv("H12")
        accrual_amount  = rv("H13")
        total_accrual   = rv("H16")
        final_annual    = rv("M13")
        final_total     = rv("M16")
        accumulation    = rv("G13")
        raw_deduction   = rv("C22")
        delayed_payment = rv("G14")

        # Преобразуем raw_deduction (строка или число) в целое число рублей
        if isinstance(raw_deduction, (int, float)):
            deduction_rub = int(raw_deduction)
        else:
            s = str(raw_deduction).lower()
            s = s.replace("рублей", "").replace("рубль", "").strip()
            s = s.replace(" ", "").replace(",", ".")
            try:
                val = float(s)
                deduction_rub = int(val)
            except:
                deduction_rub = 0

        wb.Close(SaveChanges=False)
        excel.Quit()
        pythoncom.CoUninitialize()

        result = {
            "contribution":       annual_sum,
            "totalContribution":  annual_sum * term,
            "accrualPercent":     accrual_percent,
            "accrualAmount":      accrual_amount,
            "totalAccrual":       total_accrual,
            "finalAnnual":        final_annual,
            "finalTotal":         final_total,
            "accumulation":       accumulation,
            "totalDeduction":     deduction_rub,
            "delayedPayment":     delayed_payment
        }

        print("[Server] CAREFUTURE: Возвращаем результат:", result)
        return jsonify(result), 200

    except Exception as e:
        print("[Server] CAREFUTURE: ОШИБКА в целом:", e)
        try:
            excel.Quit()
            pythoncom.CoUninitialize()
        except:
            pass
        return jsonify({"error": str(e)}), 500


@app.route('/api/proxy/carefuture/save', methods=["POST", "OPTIONS"])
def proxy_carefuture_save():
    if request.method == "OPTIONS":
        return cors_preflight_response()
    data = request.get_json(force=True)
    try:
        target = url_for('carefuture_save', _external=True)
        resp = requests.post(
            target,
            json=data,
            headers={'Content-Type': 'application/json'}
        )
        return Response(
            resp.content,
            status=resp.status_code,
            content_type=resp.headers.get('Content-Type')
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/carefuture/send_manager', methods=["POST", "OPTIONS"])
def carefuture_send_manager():
    if request.method == "OPTIONS":
        return cors_preflight_response()
    data = request.get_json()
    try:
        subject = data.get("subject", "")
        body    = data.get("body", "")

        smtp_server = "smtp.yandex.ru"
        smtp_port   = 587
        smtp_user   = "rgszh-miniapp@yandex.ru"
        smtp_pass   = "ofuhaiuymneawcve"
        recipients  = ["zerotlt@mail.ru", "I.dav@mail.ru"]

        message = (
            f"From: {smtp_user}\r\n"
            f"To: {', '.join(recipients)}\r\n"
            f"Subject: {subject}\r\n"
            "Content-Type: text/plain; charset=utf-8\r\n\r\n"
            f"{body}"
        )

        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_user, recipients, message.encode('utf-8'))
        server.quit()

        return jsonify({"status": "ok"}), 200

    except Exception as e:
        print("[Server] Ошибка при отправке менеджеру CAREFUTURE:", e)
        return jsonify({"error": str(e)}), 500

@app.route('/api/proxy/carefuture/send_manager', methods=["POST", "OPTIONS"])
def proxy_carefuture_send_manager():
    if request.method == "OPTIONS":
        return cors_preflight_response()
    data = request.get_json(force=True)
    try:
        target = url_for('carefuture_send_manager', _external=True)
        resp = requests.post(
            target,
            json=data,
            headers={'Content-Type': 'application/json'}
        )
        return Response(
            resp.content,
            status=resp.status_code,
            content_type=resp.headers.get('Content-Type')
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ===== Раздача статических файлов React =====

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_frontend(path):
    if path and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    # Иначе — возвращаем index.html
    return send_from_directory(app.static_folder, 'index.html')


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)














